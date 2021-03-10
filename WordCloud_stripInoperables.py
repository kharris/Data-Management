import openpyxl, pathlib, pandas as pd
from pathlib import Path
from openpyxl import Workbook
from sklearn.feature_extraction.text import ENGLISH_STOP_WORDS

#Set the path object for the input excel spreadsheet and import it into a pandas dataframe
origDataPath = Path(r"C:\Documents\Projects\Data Analytics\Procurement_Report.xlsx")
origDatadf = pd.read_excel(origDataPath, sheet_name='Procurement_Report__All')

#Initialize empty lists to record articles, prepositions, and operational words found in the data
articleList = []
prepList = []
operWordset = []

#Assign records in the Procurement Description field with a corresponding Type = Design and Construction to a new series 
possWords = origDatadf['Procurement Description'].loc[origDatadf['Type of Procurement'] == 'Design and Construction/Maintenance']

#Loop through the new series of records and collect in the operWordset list all words that are more than 2 characters and are not in the "ENGLISH_STOP_WORDS" set imported from sklearn
for x in possWords:
    operWordset.extend([y.lower() for y in filter(lambda f: (f.isalpha()) and (f.lower() not in ENGLISH_STOP_WORDS) and (len(f) >= 2), x.split(' '))])

#Open the excel workbook containing the input data and create a new spreadsheet called "ProjNameWords"
wb = openpyxl.load_workbook(str(origDataPath))
ws = wb.create_sheet(title='ProjNameWords')

#Place the Spreadsheet title in the first cell and list each word in operWordset in a different row in column A
ws['A1'].value = 'ProjNameWords'
for n,x in enumerate(operWordset,2):
    ws[f'A{n}'].value = x

#Save and close the excel spreadsheet
wb.save(r"C:\Documents\Projects\Data Analytics\Procurement_Report.xlsx")
wb.close()
print('Finished')
